from typing import Optional
from urllib.parse import urlsplit, unquote
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import openpyxl
from openpyxl.styles import Alignment, Font
import logging
import tomli


logger = logging.getLogger('exporters')
logging.basicConfig(filename='/tmp/api.censusreporter.org.error.log',level=logging.DEBUG)

Session = sessionmaker()


with open('config.toml', 'rb') as f:
    config = tomli.load(f)


_sessions = {}
def session(sql_url):
    try:
        return _sessions[sql_url]
    except KeyError: # probably not super thread-safe, but repeated execution should be harmless
        engine = create_engine(sql_url)
        _sessions[sql_url] = Session(bind=engine.connect())
        return _sessions[sql_url]


def get_sql_config(*_):
    """Return a tuple of strings: (host, user, password, database)"""

    return (
        config['db']['host'],
        config['db']['username'],
        unquote(config['db']['password']),
        config['db']['name'],
    )


def noneproof_divide(a: Optional[float], b: Optional[float]) -> Optional[float]:
    if (a is None) or (b is None):
        return None
    return a / b


def create_excel_download(sql_url, data, table_metadata, valid_geo_ids, file_ident, out_filename, format, logger=logger):
    def excel_helper(sheet, table_id, table, option):
        """
        Create excel sheet.

        :param option: 'value' or 'percent'
        """

        # Write table id and title on the first row
        sheet['A1'] = table_id
        sheet['B1'] = table['title']

        sheet['A1'].font = Font(bold=True)
        sheet['B1'].font = Font(bold=True)

        header = []
        max_indent = 0
        # get column headers
        for column_id, column_info in table['columns'].items():
            column_name_utf8 = str(column_info['name'])
            indent = column_info['indent']

            header.append((column_name_utf8, indent))

            if indent > max_indent:
                max_indent = indent

        # Populate first column with headers
        for i, col_tuple in enumerate(header):
            header = col_tuple[0]
            current_row = i + 4 # 1-based index, account for geographic headers
            current_cell = sheet.cell(row=current_row, column=1)
            current_cell.value = header
            indent = col_tuple[1]
            if indent is None:
                logger.warn("Null indent for {} {}".format(table_id, header))
                indent = 0
            current_cell.alignment = Alignment(indent=indent, wrap_text=True)

        # Resize column width
        sheet.column_dimensions['A'].width = 50

        # this SQL echoed in OGR export but no geom so copying instead of factoring out
        # plus different binding when using SQLAlchemy
        result = session(sql_url).execute(
            """SELECT full_geoid,display_name
                     FROM tiger2021.census_name_lookup
                     WHERE full_geoid IN :geoids
                     ORDER BY full_geoid""",
            {'geoids': tuple(valid_geo_ids)}
        )

        geo_headers = []
        any_zero_denominators = False
        has_denominator_column = False
        for i, (geoid, name) in enumerate(result):
            geo_headers.append(name)
            col_values = []
            col_errors = []
            for (table_id, table) in table_metadata.items():
                table_estimates = data[geoid][table_id]['estimate']
                table_errors = data[geoid][table_id]['error']
                if option == 'value':
                    for column_id, column_info in table['columns'].items():
                        col_values.append(table_estimates[column_id])
                        col_errors.append(table_errors[column_id])
                elif option == 'percent':
                    if table['denominator_column_id'] is not None:
                        has_denominator_column = True
                        base_estimate = data[geoid][table_id]['estimate'][table['denominator_column_id']]

                        for column_id, column_info in table['columns'].items():
                            if base_estimate is not None and base_estimate != 0:
                                col_values.append(noneproof_divide(table_estimates[column_id], base_estimate))
                                col_errors.append(noneproof_divide(table_errors[column_id], base_estimate))
                            else:
                                any_zero_denominators = True
                                col_values.append('*')
                                col_errors.append('')
                    else:
                        col_values.append('*')
                        col_errors.append('')


            for j, value in enumerate(col_values):
                col_num = (i + 1) * 2
                row_num = j + 4
                sheet.cell(row=row_num, column=col_num).value = value
                sheet.cell(row=row_num, column=col_num + 1).value = col_errors[j]
                if option == 'percent':
                    sheet.cell(row=row_num, column=col_num).number_format = '0.00%'
                    sheet.cell(row=row_num, column=col_num + 1).number_format = '0.00%'
            
        if option == 'percent' and (any_zero_denominators or not has_denominator_column):
            annotation_cell = sheet.cell(row=(row_num + 1), column=1)
            annotation_cell.font = Font(italic=True)
            if any_zero_denominators:
                annotation_cell.value = "* Base value of zero; no percentage available"
            elif not has_denominator_column:
                annotation_cell.value = "* Percentage values not appropriate for this table"
            else:
                annotation_cell.value = "* Unexpected error. Please contact Census Reporter at https://censusreporter.uservoice.com/ and let us know the page from where you downloaded this data."

        # Write geo headers
            current_cell = sheet.cell(row=2, column=current_col)
            current_cell.value = geo_headers[i]
            current_cell.alignment = Alignment(horizontal='center')
            sheet.merge_cells(start_row=2, end_row=2, start_column=current_col, end_column=current_col + 1)
            sheet.cell(row=3, column=current_col).value = "Value"
            sheet.cell(row=3, column=current_col + 1).value = "Error"

    wb = openpyxl.workbook.Workbook()

    # For every table in table_metadata, make a two sheets (values and percentages)
    for i, (table_id, table) in enumerate(table_metadata.items()):
        sheet = wb.active
        sheet.title = table_id + ' Values'
        excel_helper(sheet, table_id, table, 'value')

        sheet_percents = wb.create_sheet(table_id + ' Percentages')
        excel_helper(sheet_percents, table_id, table, 'percent')

    wb.save(out_filename)

def create_ogr_download(sql_url, data, table_metadata, valid_geo_ids, file_ident, out_filename, format, logger=logger):
    from osgeo import ogr
    from osgeo import osr
    format_info = supported_formats[format]
    driver_name = format_info['driver']
    ogr.UseExceptions()
    in_driver = ogr.GetDriverByName("PostgreSQL")

    host, user, password, database = get_sql_config(sql_url) 

    port = '5432'
    conn = in_driver.Open(f"PG: host=localhost dbname={database} user={user} password={password} port={port}")

    if conn is None:
        raise Exception("Could not connect to database to generate download.")


    out_driver = ogr.GetDriverByName(driver_name)
    out_srs = osr.SpatialReference()
    out_srs.ImportFromEPSG(4326)
    out_data = out_driver.CreateDataSource(out_filename)
    # See http://gis.stackexchange.com/questions/53920/ogr-createlayer-returns-typeerror
    out_layer = out_data.CreateLayer(str(file_ident), srs=out_srs, geom_type=ogr.wkbMultiPolygon)
    out_layer.CreateField(ogr.FieldDefn('geoid', ogr.OFTString))
    out_layer.CreateField(ogr.FieldDefn('name', ogr.OFTString))
    for (table_id, table) in table_metadata.items():
        for column_id, column_info in table['columns'].items():
            column_name_utf8 = str(column_id)
            if driver_name == "ESRI Shapefile":
                # Work around the Shapefile column name length limits
                out_layer.CreateField(ogr.FieldDefn(column_name_utf8, ogr.OFTReal))
                out_layer.CreateField(ogr.FieldDefn(column_name_utf8 + "e", ogr.OFTReal))
            else:
                out_layer.CreateField(ogr.FieldDefn(column_name_utf8, ogr.OFTReal))
                out_layer.CreateField(ogr.FieldDefn(column_name_utf8 + ", Error", ogr.OFTReal))

    # this SQL echoed in Excel export but no geom so copying instead of factoring out
    sql = """SELECT geom,full_geoid,display_name
             FROM tiger2021.census_name_lookup
             WHERE full_geoid IN (%s)
             ORDER BY full_geoid""" % ', '.join("'%s'" % str(g) for g in valid_geo_ids)
    in_layer = conn.ExecuteSQL(sql)

    in_feat = in_layer.GetNextFeature()
    while in_feat is not None:
        out_feat = ogr.Feature(out_layer.GetLayerDefn())
        if format in ('shp', 'kml', 'geojson'):
            out_feat.SetGeometry(in_feat.GetGeometryRef())
        geoid = in_feat.GetField('full_geoid')
        out_feat.SetField('geoid', geoid)
        out_feat.SetField('name', in_feat.GetField('display_name'))
        for (table_id, table) in table_metadata.items():
            table_estimates = data[geoid][table_id]['estimate']
            table_errors = data[geoid][table_id]['error']
            for column_id, column_info in table['columns'].items():
                column_name_utf8 = str(column_id)
                if column_id in table_estimates:
                    if format == 'shp':
                        # Work around the Shapefile column name length limits
                        estimate_col_name = column_name_utf8
                        error_col_name = column_name_utf8 + "e"
                    else:
                        estimate_col_name = column_name_utf8
                        error_col_name = column_name_utf8 + ", Error"

                    out_feat.SetField(estimate_col_name, table_estimates[column_id])
                    out_feat.SetField(error_col_name, table_errors[column_id])

        out_layer.CreateFeature(out_feat)
        in_feat.Destroy()
        in_feat = in_layer.GetNextFeature()
    out_data.Destroy()

supported_formats = { # these should all have a 'function' with the right signature
    'shp':      {"function": create_ogr_download, "driver": "ESRI Shapefile"},
    'kml':      {"function": create_ogr_download, "driver": "KML"},
    'geojson':  {"function": create_ogr_download, "driver": "GeoJSON"},
    'xlsx':     {"function": create_excel_download, "driver": "XLSX"},
    'csv':      {"function": create_ogr_download, "driver": "CSV"},
}
